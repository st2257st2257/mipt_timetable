# Скрипт превращает расписание в финальную версию с заполненными
# полями. Поля объединены для удобства использования
# (c) Aleksandr Kristal @st2257

import openpyxl
from openpyxl import load_workbook

wb = load_workbook(filename='data6.xlsx')
sheet_ranges = wb['6 курс']


# считает число букв в строке
def count_letters(text):
    count = 0
    for char in text:
        if char.isalpha():
            count += 1
    return count


# выдаёт все ячейки из диапазона
def get_all_cells(my_range):
    # в строку для функции слияния
    my_string = str(my_range)
    # разъединяем ячейки по выбранному диапазону
    sheet_ranges.unmerge_cells(my_string)

    # разделяем диапазон на начало и конец
    start_cell, end_cell = my_string.split(':')

    # считаем в интеджерах начало и конец
    start_row = int(start_cell[count_letters(start_cell):])
    start_col = openpyxl.utils.column_index_from_string(start_cell[:count_letters(start_cell)])
    end_row = int(end_cell[count_letters(end_cell):])
    end_col = openpyxl.utils.column_index_from_string(end_cell[:count_letters(end_cell)])

    values = []
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            # проходимся по всем вариантам и добавляем в массив
            cell = sheet_ranges.cell(row=row, column=col)
            values.append(cell)
    return values


# получаем все слитые ячейки
for my_range in list(sheet_ranges.merged_cells.ranges):
    # получаем первую ячейку диапазона и запоминаем её значение в VALUE
    first_cell = sheet_ranges.cell(row=my_range.min_row, column=my_range.min_col)
    value = first_cell.value

    # получаем все клетки диапазона
    range_array = get_all_cells(my_range)
    for item in range_array:
        print(sheet_ranges.cell(
            row=item.row,
            column=item.column))
        # меняем значение на основное значение в диапазоне
        sheet_ranges.cell(
            row=item.row,
            column=item.column).value = value

wb.save("res_data6.xlsx")
