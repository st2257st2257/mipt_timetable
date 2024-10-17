import openpyxl
from openpyxl import load_workbook

wb = load_workbook(filename='data.xlsx')
sheet_ranges = wb['1 курс']


def count_letters(text):
    count = 0
    for char in text:
        if char.isalpha():
            count += 1
    return count


def get_all_cells(my_range):
    print(str(my_range))
    my_string = str(my_range)
    sheet_ranges.unmerge_cells(my_string)

    # Разделяем строку на начальную и конечную ячейки
    start_cell, end_cell = my_string.split(':')

    # Получаем координаты начальной и конечной ячеек
    start_row = int(start_cell[count_letters(start_cell):])
    start_col = openpyxl.utils.column_index_from_string(start_cell[:count_letters(start_cell)])
    end_row = int(end_cell[count_letters(end_cell):])
    end_col = openpyxl.utils.column_index_from_string(end_cell[:count_letters(end_cell)])
    print(start_row, start_col, end_col, end_row)

    # Создаем список для хранения значений ячеек
    values = []

    # Перебираем ячейки в диапазоне
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            print('-')
            cell = sheet_ranges.cell(row=row, column=col)
            print('-')
            print(cell.row, cell.column)
            #sheet_ranges.unmerge_cells(
            #    start_row=cell.row,
            #    start_column=cell.column,
            #    end_row=cell.row,
            #    end_column=cell.column)
            print('-')
            values.append(cell)
            print(len(values))
    return values


for my_range in list(sheet_ranges.merged_cells.ranges):
    first_cell = sheet_ranges.cell(row=my_range.min_row, column=my_range.min_col)
    value = first_cell.value

    range_array = get_all_cells(my_range)
    # sheet_ranges.unmerge_cells(str(my_range))
    for item in range_array:
        print(sheet_ranges.cell(
            row=item.row,
            column=item.column))
        sheet_ranges.cell(
            row=item.row,
            column=item.column).value = value
        # item.value = value

    # sheet_ranges.unmerge_cells(str(my_range))

wb.save("res_data.xlsx")
